#!/usr/bin/env python3
"""Final expanded fixed-snapshot 2013 corpus builder."""
from __future__ import annotations

import csv
import json
import re
import shutil
from pathlib import Path
from urllib.parse import quote

from bs4 import BeautifulSoup
from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

import build_corpus_2013_v5 as base
import build_corpus_2013_v7 as v7

YIHUI_REPO = "https://github.com/yihui/yihui.org.git"
YIHUI_COMMIT = "d660c32dcee301deeabe2714d02b3402b15f99c4"
GOVDATA_REPO = "https://github.com/ChengZhiFU609/DataMining-InformationExtraction_and_Sentiment_Analysis_of_Government_Work_Reports.git"
GOVDATA_COMMIT = "94306f5562e9fb9d541350521582953b2bd65134"


def clean_markdown_article(raw: str) -> tuple[str, str]:
    title = ""
    if raw.startswith("---"):
        match = re.match(r"---\s*\n(.*?)\n---\s*\n", raw, flags=re.S)
        if match:
            header = match.group(1)
            title_match = re.search(r"^title:\s*[\"']?(.*?)[\"']?\s*$", header, flags=re.M)
            if title_match:
                title = base.CC.convert(title_match.group(1).strip())
            raw = raw[match.end():]
    raw = re.sub(r"```.*?```", "", raw, flags=re.S)
    raw = re.sub(r"`[^`]+`", "", raw)
    html = __import__("markdown").markdown(raw, extensions=["extra"])
    soup = BeautifulSoup(html, "lxml")
    for selector in "script style img figure figcaption table pre code blockquote".split():
        for node in soup.select(selector):
            node.decompose()
    lines = [e.get_text(" ", strip=True) for e in soup.find_all(["h2", "h3", "h4", "p", "li"])]
    if not title:
        h = soup.find(["h1", "h2"])
        title = base.CC.convert(h.get_text(" ", strip=True)) if h else "2013年中文博客评论"
        if lines and lines[0] == title:
            lines.pop(0)
    return title, base.clean_paragraphs(lines)


def collect_yihui(pool: list[dict], repo: Path) -> None:
    for path in sorted((repo / "content/cn").glob("2013-*.md")):
        raw = path.read_text("utf-8", errors="ignore")
        title, body = clean_markdown_article(raw)
        date_match = re.match(r"(2013-\d{2}-\d{2})", path.name)
        date = date_match.group(1) if date_match else "2013"
        official = "https://yihui.org/cn/" + path.stem + "/"
        source = (
            "https://github.com/yihui/yihui.org/blob/"
            f"{YIHUI_COMMIT}/" + quote(str(path.relative_to(repo)), safe="/")
        )
        base.add(
            pool,
            title=title,
            author="谢益辉",
            date=date,
            source_url=source,
            official_url=official,
            platform="谢益辉中文博客固定仓库快照",
            genre="个人观点 / 开源与技术文化评论",
            rights="MIT License (repository content and documentation)",
            license_url="https://github.com/yihui/yihui.org/blob/d660c32dcee301deeabe2714d02b3402b15f99c4/LICENSE",
            text=body,
            notes=f"Fixed repository commit; official_url={official}; front matter, code, images, block quotations and link markup removed; main Chinese prose retained.",
        )
    print("YIHUI", len([x for x in pool if x["source_platform"] == "谢益辉中文博客固定仓库快照"]))


def read_candidate(path: Path) -> str:
    suffix = path.suffix.lower()
    try:
        if suffix in {".txt", ".md", ".rst", ".csv", ".json"}:
            return path.read_text("utf-8", errors="ignore")
        if suffix == ".docx":
            return "\n".join(p.text for p in Document(path).paragraphs)
        if suffix == ".xlsx":
            wb = load_workbook(path, read_only=True, data_only=True)
            values = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    values.extend(str(v) for v in row if v is not None)
            return "\n".join(values)
        if suffix == ".pdf":
            return "\n".join(page.extract_text() or "" for page in PdfReader(path).pages)
    except Exception as exc:
        print("READ_FAIL", path, exc)
    return ""


def report_title(raw: str, path: Path) -> str:
    for line in raw.splitlines()[:80]:
        value = base.normalize_line(line)
        if "2013" in value and "政府工作报告" in value and len(value) < 100:
            return value
        if "政府工作报告" in value and len(value) < 80:
            return "2013年" + value if "2013" not in value else value
    stem = base.CC.convert(path.stem)
    return stem if "2013" in stem else "2013年" + stem


def clean_official_report(raw: str) -> str:
    start_positions = [raw.find(key) for key in ("各位代表", "过去五年", "一、") if raw.find(key) >= 0]
    if start_positions:
        raw = raw[min(start_positions):]
    raw = re.split(r"参考文献|数据来源|附录", raw)[0]
    lines = []
    for line in raw.splitlines():
        value = line.strip()
        if not value:
            continue
        if re.match(r"^(?:标题|作者|来源|发布时间|发布日期)[:：]", value):
            continue
        lines.append(value)
    return base.clean_paragraphs(lines)


def collect_government_dataset(pool: list[dict], repo: Path) -> None:
    accepted = 0
    for path in sorted(repo.rglob("*")):
        if not path.is_file() or "2013" not in str(path):
            continue
        if path.suffix.lower() not in {".txt", ".md", ".rst", ".csv", ".json", ".docx", ".xlsx", ".pdf"}:
            continue
        if path.stat().st_size > 12_000_000:
            continue
        raw = read_candidate(path)
        if "政府工作报告" not in raw and "各位代表" not in raw:
            continue
        body = clean_official_report(raw)
        if base.nows(body) < 1200:
            continue
        title = report_title(raw, path)
        relative = quote(str(path.relative_to(repo)), safe="/")
        source = f"https://github.com/ChengZhiFU609/DataMining-InformationExtraction_and_Sentiment_Analysis_of_Government_Work_Reports/blob/{GOVDATA_COMMIT}/{relative}"
        base.add(
            pool,
            title=title,
            author="相关地方人民政府 / 国家机关",
            date="2013",
            source_url=source,
            official_url="https://www.gov.cn/",
            platform="政府工作报告数据仓库固定快照",
            genre="地方治理回顾 / 政策评估 / 政府工作报告",
            rights="Public domain: PRC Copyright Law Article 5 official document",
            license_url=base.PUBLIC_LICENSE_URL,
            text=body,
            notes="Official government report found in a fixed government-report data-mining repository snapshot; main report text only; metadata and appendices removed.",
        )
        accepted += 1
    print("GOVDATA", accepted)


def main() -> None:
    shutil.rmtree(base.OUT, ignore_errors=True)
    root = Path("/tmp/corpus2013_sources")
    shutil.rmtree(root, ignore_errors=True)
    root.mkdir(parents=True)
    ruanyf = root / "read"
    coolshell = root / "haoel"
    gov = root / "gov"
    yihui = root / "yihui"
    govdata = root / "govdata"
    base.clone_at(base.RUANYF_REPO, base.RUANYF_COMMIT, ruanyf)
    base.clone_at(base.COOLSHELL_REPO, base.COOLSHELL_COMMIT, coolshell)
    base.clone_at(base.GOV_REPO, base.GOV_COMMIT, gov)
    base.clone_at(YIHUI_REPO, YIHUI_COMMIT, yihui)
    base.clone_at(GOVDATA_REPO, GOVDATA_COMMIT, govdata)
    pool: list[dict] = []
    v7.collect_all_chinese_ruanyf(pool, ruanyf)
    base.collect_coolshell(pool, coolshell)
    base.collect_government(pool, gov)
    collect_yihui(pool, yihui)
    collect_government_dataset(pool, govdata)
    total = sum(x["char_count"] for x in pool)
    print("POOL", len(pool), total, sorted({x["source_platform"] for x in pool}))
    if total < 225_000:
        raise SystemExit(f"insufficient pool: {total}")
    bins, unused = base.choose_bins(pool)
    print("BIN_RAW", [sum(x["char_count"] for x in selected) for selected in bins])
    base.write_outputs(bins, unused)


if __name__ == "__main__":
    main()
